import pandas as pd
import os
import re
import openpyxl


excludedKeywords = {"trade", "tax", "time deposit", "restricted", "property", "prepayment", "prepaid", "project", "intangible", "inventor", 'advance', "cash", "fixed asset", "oil", "gas",
                    "accounts", "unused", "stockpile", "stripping", "cost", "right", "develop", "industrial", "impairment", "goodwill", "pre-paid", "expense", "depreciation", "benefit", "deferred",
                    "current asset", "(december 31, 2013:us8", "886 as of december 31, 2004 and 2005, respectively", "account receivable", "amounts due from related parties", 
                    "another s financial assets", "bearer plant", "breeding", "cattle", "current maturities", "acquisition", "contract asset", "december",
                    "net of accumulated other assets amortization of rp", "propert", "gross receivable", "unbilled receivable", "receivable from", "receivables from", "rp", "retention", "r.p"}


includedKeywords = {"forsale", "forsale", "financialasset", "contractasset", "swap", "financialinstrument", "december", "derivative","acquisition","dividend","due", "factoring", "financelease",
              "option", "forward", "investment","joint", "loanto", "loansto", "longterm", "mutual", "receiavbles", "otherreceivable", "others", "receivablesfrom", "receivablefrom",
              "receivables-other","relatedparties","rp", "thirdpart"}

includedKeywords2 = {"receivable", "securities", "investment", "forsale", "security", "derivative", "sale"}
includedKeywords3 = {"totalreceivables", "accountsreceivablelongterm", "loansreceivablelongterm", "longterminvestments", "shortterminvestments", "tradingassetsecurities", "totalassets"}
bumn = [
    'IDX:BMRI',
    'IDX:BBNI',
    'IDX:BBRI',
    'IDX:BBTN',
    'IDX:BRIS',
    'IDX:BJBR',
    'IDX:BEKS',
    'IDX:BJTM',
    'IDX:PGAS',
    'IDX:PTBA',
    'IDX:ELSA',
    'IDX:TLKM',
    'IDX:MTEL',
    'IDX:JSMR',
    'IDX:PTPP',
    'IDX:PPRO',
    'IDX:WIKA',
    'IDX:ADHI',
    'IDX:TINS',
    'IDX:ANTM',
    'IDX:SMGR',
    'IDX:SMBR',
    'IDX:KRAS',
    'IDX:WSBP',
    'IDX:KAEF',
    'IDX:INAF',
]

asset_keys = {}
yearHeaderIdentifier = "Recommended: S&P Capital IQ - Standard"
header = ["Fiscal Quarter"]
numOfQuarters = 59 #number of quarter from 2010 to 2024

def list_files_in_current_folder():
    """Lists all files in the current folder."""
    try:
        absolute_path = os.path.abspath('KOMPAS100 Report')
        files = [
            os.path.join('KOMPAS100 Report', f)
            for f in os.listdir(absolute_path)
            if os.path.isfile(os.path.join(absolute_path, f)) and (f.endswith(".xls") or f.endswith(".xlsx"))
        ]  
        return files
    except OSError as e:
        print(f"Error accessing directory: {e}")

def containsExcludedKeyword(account):
    for keyword in excludedKeywords:
        if keyword in account.lower():
            return True
    return False

def containsIncludedKeyword(account: str):
    for keyword in includedKeywords3:
        parsedAccount = re.sub(r'[^a-zA-Z0-9]', '', account.lower())
        if keyword == parsedAccount:
            return True
    return False

def renameFile(old_filename, new_filename):
    """Renames a file.

    Args:
        old_filename: The current filename (including path if necessary).
        new_filename: The desired new filename (including path if necessary).
    """
    try:
        os.rename(old_filename, new_filename)
        print(f"File '{old_filename}' renamed to '{new_filename}'")
    except FileNotFoundError:
        print(f"Error: File '{old_filename}' not found.")
    except FileExistsError:
        print(f"Error: File '{new_filename}' already exists.")
    except Exception as e:
        print(f"An error occurred: {e}")

def listAccounts():
    excel_files = list_files_in_current_folder()
    for file in excel_files:
        if file.endswith(".xls") or file.endswith(".xlsx") and "output" not in file:
            df = pd.read_excel(file, sheet_name="Balance Sheet")
            ticker = df.iloc[1, 0].split(" (MI KEY")[0]
            start_print = False
            for i in range(0, len(df)):
                if "asset" in str(df.iloc[i, 0]).lower():
                    start_print = True
                if "liabilit" in str(df.iloc[i,0]).lower():
                    break
                if start_print == True and containsIncludedKeyword(df.iloc[i,0]):
                    if df.iloc[i, 0] in asset_keys:
                        asset_keys[df.iloc[i, 0]].append(ticker)
                    else:
                        asset_keys[df.iloc[i, 0]] = [ticker]

def add_row_from_array(sheet, row_data, skipRow = 1):
    row_num = sheet.max_row + skipRow
    for col_num, cell_value in enumerate(row_data, 1):
        sheet.cell(row=row_num, column=col_num, value=cell_value)
        
def add_column_from_array(sheet, col_data):
    col_num = sheet.max_column + 1
    for row_num, cell_value in enumerate(col_data, 1):
        sheet.cell(row=row_num, column=col_num, value=cell_value)
        
        
def writeAllAssetsToFormattedExcel(workbook: openpyxl.Workbook, excelFiles: list):
    header = []
    data = []
    for file in sorted(excelFiles):
        if file.endswith(".xls") or file.endswith(".xlsx") and "output" not in file:
            df = pd.read_excel(file, sheet_name="Balance Sheet")
            df.fillna(0, inplace=True)
            ticker = df.iloc[1, 0].split(" (MI KEY")[0]
            data.append([ticker])
            for i in range(0, len(df)):
                identifier = str(df.iloc[i, 0])
                if identifier == yearHeaderIdentifier and len(header) == 0:
                    for j in range(0, len(df.iloc[i])):
                        header.append(df.iloc[i,j])                    
                if "liabilit" in identifier.lower():
                    break
                if identifier in asset_keys:
                    asset_keys[df.iloc[i, 0]].append(ticker)
                    rowData = []
                    for j in range(0, len(df.iloc[i])):
                        if pd.isna(df.iloc[i,j]):
                            rowData.append(0)
                        else:
                            rowData.append(df.iloc[i,j])
                    data.append(rowData)

    sheet1 = workbook.active
    sheet1.title = "Balance Sheet"
    add_row_from_array(sheet1, header, 1)
    for i in data:
        if len(i) == 1:
            add_row_from_array(sheet1, i, 3)
        else:
            add_row_from_array(sheet1, i, 1)
        
def writeTotalAssetsToFormattedExcel(workbook: openpyxl.Workbook, excelFiles: list):
    tickerToTotalAssetDict = {}
    for file in excelFiles:
        if file.endswith(".xls") or file.endswith(".xlsx") and "output" not in file:
            df = pd.read_excel(file, sheet_name="Balance Sheet")
            df.fillna(0, inplace=True)
            ticker = df.iloc[1, 0].split(" (MI KEY")[0]
        for i in range(0, len(df)):
            identifier = str(df.iloc[i, 0])
            if identifier == yearHeaderIdentifier and len(header) == 1:
                for j in range(1, len(df.iloc[i])):
                    header.append(df.iloc[i,j])
            if identifier == 'Total Assets':
                columnData = [ticker]
                for j in range(1, len(df.iloc[i])):
                    columnData.append(df.iloc[i,j])
                tickerToTotalAssetDict[ticker] = columnData
    sheet = workbook.create_sheet("Total Asset")
    add_column_from_array(sheet, header)
    for ticker in sorted(tickerToTotalAssetDict):
        add_column_from_array(sheet, tickerToTotalAssetDict[ticker])
                
def writeFinancialAssetsToFormattedExcel(workbook: openpyxl.Workbook, excelFiles):
    tickerToTotalAssetDict = {}
    for file in sorted(excelFiles):
        if file.endswith(".xls") or file.endswith(".xlsx") and "output" not in file:
            df = pd.read_excel(file, sheet_name="Balance Sheet")
            df.fillna(0, inplace=True)
            ticker = df.iloc[1, 0].split(" (MI KEY")[0]
        for i in range(0, len(df)):
            identifier = str(df.iloc[i, 0])
            if identifier == yearHeaderIdentifier and len(header) == 1:
                for j in range(1, len(df.iloc[i])):
                    header.append(df.iloc[i,j])
            if identifier == 'Total Assets':
                break
            if identifier in asset_keys:
                columnData = tickerToTotalAssetDict[ticker] if ticker in tickerToTotalAssetDict else [ticker]
                if len(columnData) == 1:
                    for j in range(1, len(df.iloc[i])):
                        columnData.append(df.iloc[i,j] if not pd.isna(df.iloc[i,j]) else 0)
                else:
                    for j in range(1, len(df.iloc[i])):
                        columnData[j] = columnData[j] + (df.iloc[i,j] if not pd.isna(df.iloc[i,j]) else 0)
                tickerToTotalAssetDict[ticker] = columnData
    sheet = workbook.create_sheet("Financial Asset")
    add_column_from_array(sheet, header)
    for ticker in sorted(tickerToTotalAssetDict):
        add_column_from_array(sheet, tickerToTotalAssetDict[ticker])
        
def writeInstitutionalOwnershipHistoryToFormattedExcel(workbook: openpyxl.Workbook, excelFiles, numOfQuarters):
    tickerToInstitutionalOwnershipHistory = {}
    for file in sorted(excelFiles):
        isStartWrite = False
        if file.endswith(".xls") or file.endswith(".xlsx") and "output" not in file:
            df = pd.read_excel(file, sheet_name="Ownership History")
            df.fillna(0, inplace=True)
            ticker = df.iloc[1, 0].split(" (MI KEY")[0]
        tickerToInstitutionalOwnershipHistory[ticker] = [ticker] + [0] * numOfQuarters
        for i in range(0, len(df)):
            if df.iloc[i, 0] == 'Holder':
                isStartWrite = True
            elif isStartWrite:
                columnData = tickerToInstitutionalOwnershipHistory[ticker]
                temp = 1
                for j in range(len(df.iloc[i])-3, 1, -1):
                    columnData[temp] = columnData[temp] + float(df.iloc[i,j])
                    temp += 1
                tickerToInstitutionalOwnershipHistory[ticker] = columnData
    sheet = workbook.create_sheet("Institutional Ownership")
    add_column_from_array(sheet, header)
    for ticker in sorted(tickerToInstitutionalOwnershipHistory):
        add_column_from_array(sheet, tickerToInstitutionalOwnershipHistory[ticker])



# UNCOMMENT this to rename the file names that are generated from CapitalIQ
excel_files = list_files_in_current_folder()
# for file in sorted(excel_files):
#     print(file)
#     df = pd.read_excel(file, sheet_name="Balance Sheet")
#     ticker = df.iloc[1, 0].split(" (MI KEY")[0]
#     renameFile(file, "KOMPAS100 Report/{}.xls".format(ticker))

workbook = openpyxl.Workbook()
listAccounts()
writeAllAssetsToFormattedExcel(workbook, excel_files)
writeTotalAssetsToFormattedExcel(workbook, excel_files)
writeFinancialAssetsToFormattedExcel(workbook, excel_files)
writeInstitutionalOwnershipHistoryToFormattedExcel(workbook, excel_files, numOfQuarters)
workbook.save("output.xlsx")